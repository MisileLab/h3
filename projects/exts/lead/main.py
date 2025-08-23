import os
from datetime import datetime
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from sqlalchemy import Column, Integer, MetaData, String, Table, create_engine, select, insert, DateTime, UniqueConstraint
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.styles import Font, PatternFill, Alignment

metadata = MetaData()
response_table = Table(
  "response",
  metadata,
  Column("id", Integer, primary_key=True, autoincrement=True),
  Column("email", String),
  Column("name", String),
  Column("number", Integer),
  Column("created_at", DateTime),
  Column("updated_at", DateTime),
  Column("create_time", DateTime),
  Column("last_submitted_time", DateTime),
  UniqueConstraint('email', 'name', 'create_time', name='uq_email_name_createtime')
)

# Create an engine (adjust connection string as needed)
engine = create_engine('sqlite:///lead.db')

def get_response_by_email(email: str) -> None:
  stmt = select(response_table).where(response_table.c.email == email)
  with engine.connect() as conn:
    result = conn.execute(stmt)
    responses = result.fetchall()
    
    if not responses:
      print(f"No responses found for email: {email}")
      return
    
    print(f"Found {len(responses)} response(s) for email: {email}")
    print("=" * 60)
    
    for i, row in enumerate(responses, 1):
      print(f"Response #{i} (ID: {row.id}):")
      print(f"  Email: {row.email}")
      print(f"  Name: {row.name}")
      print(f"  Number: {row.number}")
      print(f"  Created At: {row.created_at}")
      print(f"  Updated At: {row.updated_at}")
      print(f"  Create Time: {row.create_time}")
      print(f"  Last Submitted Time: {row.last_submitted_time}")
      print("-" * 40)

def get_response_by_email_name_createtime(email: str, name: str, create_time: str) -> None:
  """Search for responses by email, name, and createTime combination"""
  try:
    # Parse the create_time string to datetime object
    parsed_time = datetime.fromisoformat(create_time.replace('Z', '+00:00'))
  except ValueError:
    print(f"Invalid create_time format: {create_time}")
    return
  
  stmt = select(response_table).where(
    response_table.c.email == email,
    response_table.c.name == name,
    response_table.c.create_time == parsed_time
  )
  
  with engine.connect() as conn:
    result = conn.execute(stmt)
    responses = result.fetchall()
    
    if not responses:
      print(f"No responses found for email: {email}, name: {name}, createTime: {create_time}")
      return
    
    print(f"Found {len(responses)} response(s) for the specified combination:")
    print("=" * 60)
    
    for i, row in enumerate(responses, 1):
      print(f"Response #{i} (ID: {row.id}):")
      print(f"  Email: {row.email}")
      print(f"  Name: {row.name}")
      print(f"  Number: {row.number}")
      print(f"  Created At: {row.created_at}")
      print(f"  Updated At: {row.updated_at}")
      print(f"  Create Time: {row.create_time}")
      print(f"  Last Submitted Time: {row.last_submitted_time}")
      print("-" * 40)

def check_duplicates() -> None:
  """Check for potential duplicate responses in the database"""
  stmt = select(
    response_table.c.email,
    response_table.c.name,
    response_table.c.create_time,
    response_table.c.id
  ).order_by(
    response_table.c.email,
    response_table.c.name,
    response_table.c.create_time
  )
  
  with engine.connect() as conn:
    result = conn.execute(stmt)
    rows = result.fetchall()
    
    if not rows:
      print("No responses found in the database.")
      return
    
    print("Checking for potential duplicates...")
    print("=" * 60)
    
    duplicates_found = False
    current_group = None
    
    for row in rows:
      group_key = (row.email, row.name, row.create_time)
      
      if group_key == current_group:
        # This is a duplicate
        if not duplicates_found:
          print("DUPLICATES FOUND:")
          duplicates_found = True
        
        print(f"  Duplicate: ID {row.id} - Email: {row.email}, Name: {row.name}, CreateTime: {row.create_time}")
      else:
        current_group = group_key
    
    if not duplicates_found:
      print("No duplicates found. All responses are unique based on email + name + createTime combination.")
    else:
      print("\nNote: Duplicates are not allowed due to unique constraint on (email, name, createTime)")

def export_to_excel(filename: str = "form_responses.xlsx") -> None:
  """Export all responses from SQLite database to Excel file"""
  try:
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    assert ws is not None, "Worksheet should not be None"
    ws.title = "Form Responses"
    
    # Define headers (removed Created At and Updated At)
    headers = [
      "ID", "Email", "Name", "Number", 
      "Create Time", "Last Submitted Time"
    ]
    
    # Style the header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add headers to the first row
    for col, header in enumerate(headers, 1):
      cell = ws.cell(row=1, column=col, value=header)
      assert cell is not None, "Cell should not be None"
      cell.font = header_font
      cell.fill = header_fill
      cell.alignment = header_alignment
    
    # Fetch all data from the database
    stmt = select(response_table)
    with engine.connect() as conn:
      result = conn.execute(stmt)
      rows = result.fetchall()
    
    if not rows:
      print("No data to export. Database is empty.")
      return
    
    # Add data rows (removed Created At and Updated At columns)
    for row_num, row in enumerate(rows, 2):
      ws.cell(row=row_num, column=1, value=row.id)
      ws.cell(row=row_num, column=2, value=row.email)
      ws.cell(row=row_num, column=3, value=row.name)
      ws.cell(row=row_num, column=4, value=row.number)
      ws.cell(row=row_num, column=5, value=row.create_time)
      ws.cell(row=row_num, column=6, value=row.last_submitted_time)
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
      max_length = 0
      
      # Get column letter safely
      try:
        column_letter = ws.cell(row=1, column=col_num).column_letter
      except AttributeError:
        # Fallback: convert column number to letter
        column_letter = chr(64 + col_num)  # A=65, B=66, etc.
      
      for row_num in range(1, len(rows) + 2):  # +2 for header row
        cell = ws.cell(row=row_num, column=col_num)
        if cell is not None and cell.value is not None:
          try:
            cell_length = len(str(cell.value))
            if cell_length > max_length:
              max_length = cell_length
          except:
            pass
      
      adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
      ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb.save(filename)
    print(f"✅ Successfully exported {len(rows)} responses to '{filename}'")
    
  except Exception as e:
    print(f"❌ Error exporting to Excel: {e}")

def get_all_responses() -> None:
  """Display all responses in the database with timestamp information"""
  stmt = select(response_table)
  with engine.connect() as conn:
    result = conn.execute(stmt)
    responses = result.fetchall()
    
    if not responses:
      print("No responses found in the database.")
      return
    
    print(f"Found {len(responses)} response(s):")
    print("=" * 60)
    
    for i, row in enumerate(responses, 1):
      print(f"Response #{i}:")
      print(f"  Email: {row.email}")
      print(f"  Name: {row.name}")
      print(f"  Number: {row.number}")
      print(f"  Created At: {row.created_at}")
      print(f"  Updated At: {row.updated_at}")
      print(f"  Create Time: {row.create_time}")
      print(f"  Last Submitted Time: {row.last_submitted_time}")
      print("-" * 40)

def extract_answers(data, field_mapping=None):
  """Extract all answers from the JSON response with optional field mapping"""
  answers = {}
  
  # Default mapping if none provided
  if field_mapping is None:
    field_mapping = {
      '7aef9547': 'name',
      '71ef93f7': 'number', 
      '61d54db6': 'email'
    }
  
  # Navigate through the JSON structure
  for response in data['responses']:
    # Extract timestamp fields
    if 'createTime' in response:
      answers['create_time'] = response['createTime']
    if 'lastSubmittedTime' in response:
      answers['last_submitted_time'] = response['lastSubmittedTime']
    
    for question_id, question_data in response['answers'].items():
      # Extract the first answer value (assuming single answers)
      answer_value = question_data['textAnswers']['answers'][0]['value']
      
      # Use mapped name if available, otherwise use original question_id
      field_name = field_mapping.get(question_id, question_id)
      print(field_name, answer_value)
      answers[field_name] = answer_value
  
  return answers

def save_responses_to_db(responses_data, field_mapping=None):
  """Extract responses and save them to the database"""
  # Create tables if they don't exist
  metadata.create_all(engine)
  
  extracted_answers = extract_answers(responses_data, field_mapping)
  
  with engine.connect() as conn:
    # Convert number to int if it exists
    if 'number' in extracted_answers:
      try:
        extracted_answers['number'] = int(extracted_answers['number'])
      except ValueError:
        extracted_answers['number'] = 0
    
    # Convert timestamp strings to datetime objects
    if 'create_time' in extracted_answers:
      try:
        extracted_answers['create_time'] = datetime.fromisoformat(
          extracted_answers['create_time'].replace('Z', '+00:00')
        )
      except ValueError:
        print(f"Warning: Could not parse create_time: {extracted_answers['create_time']}")
        extracted_answers['create_time'] = None
    
    if 'last_submitted_time' in extracted_answers:
      try:
        extracted_answers['last_submitted_time'] = datetime.fromisoformat(
          extracted_answers['last_submitted_time'].replace('Z', '+00:00')
        )
      except ValueError:
        print(f"Warning: Could not parse last_submitted_time: {extracted_answers['last_submitted_time']}")
        extracted_answers['last_submitted_time'] = None
    
    # Check if combination of email, name, and createTime already exists
    email = extracted_answers.get('email')
    name = extracted_answers.get('name')
    create_time = extracted_answers.get('create_time')
    
    if email and name and create_time:
      # Check for existing record with same email, name, and createTime
      existing = conn.execute(
        select(response_table).where(
          response_table.c.email == email,
          response_table.c.name == name,
          response_table.c.create_time == create_time
        )
      ).fetchone()
      
      if existing:
        # Update existing record
        stmt = response_table.update().where(
          response_table.c.email == email,
          response_table.c.name == name,
          response_table.c.create_time == create_time
        ).values(**extracted_answers)
        conn.execute(stmt)
        print(f"Updated response for email: {email}, name: {name}, createTime: {create_time}")
      else:
        # Insert new record
        stmt = insert(response_table).values(**extracted_answers)
        conn.execute(stmt)
        print(f"Inserted new response for email: {email}, name: {name}, createTime: {create_time}")
    else:
      # Insert without duplicate check if missing required fields
      stmt = insert(response_table).values(**extracted_answers)
      conn.execute(stmt)
      print("Inserted response without duplicate check (missing required fields)")
    
    conn.commit()

# The scope determines the level of access.
# This scope is read-only for form responses.
SCOPES = ["https://www.googleapis.com/auth/forms.responses.readonly"]

# The ID of the Google Form.
FORM_ID = input("Form ID:") 

def main():
  """
  Authenticates with the Google Forms API and prints responses from a form.
  """
  creds = None
  # The file token.json stores the user's access and refresh tokens, and is
  # created automatically when the authorization flow completes for the first
  # time.
  if os.path.exists("token.json"):
    creds = Credentials.from_authorized_user_file("token.json", SCOPES)
  # If there are no (valid) credentials available, let the user log in.
  if not creds or not creds.valid:
    if creds and creds.expired and creds.refresh_token:
      creds.refresh(Request())
    else:
      # IMPORTANT: Rename your downloaded client_secrets.json to credentials.json
      flow = InstalledAppFlow.from_client_secrets_file(
        "credentials.json", SCOPES
      )
      creds = flow.run_local_server(port=0)
      
    # Save the credentials for the next run
    with open("token.json", "w") as token:
      token.write(creds.to_json())
  try:
    # Build the service object
    service = build("forms", "v1", credentials=creds)
    # Retrieve the form responses.
    result = service.forms().responses().list(formId=FORM_ID).execute()
    print("Raw API Response:")
    print(result)
    
    # Process and save responses to database
    if 'responses' in result:
      # You may need to adjust the field mapping based on your actual question IDs
      field_mapping = {
        # Replace these with your actual question IDs from the form
        '7aef9547': 'name',
        '71ef93f7': 'number',
        '61d54db6': 'email'
      }
      
      save_responses_to_db(result, field_mapping)
      
      # Export to Excel after saving to database
      print("\n📊 Exporting data to Excel...")
      export_to_excel()
      
      # Example: Query options
      print("\nSearch options:")
      print("1. Search by email only")
      print("2. Search by email, name, and createTime combination")
      print("3. View all responses")
      print("4. Check for duplicates")
      print("5. Export to Excel")
      print("6. Skip search")
      
      choice = input("Enter your choice (1-6): ").strip()
      
      if choice == "1":
        email_to_search = input("Enter email to search: ")
        if email_to_search:
          get_response_by_email(email_to_search)
      elif choice == "2":
        email_to_search = input("Enter email: ")
        name_to_search = input("Enter name: ")
        create_time_to_search = input("Enter createTime (e.g., 2025-08-23T09:08:07.959Z): ")
        if email_to_search and name_to_search and create_time_to_search:
          get_response_by_email_name_createtime(email_to_search, name_to_search, create_time_to_search)
      elif choice == "3":
        get_all_responses()
      elif choice == "4":
        check_duplicates()
      elif choice == "5":
        export_to_excel()
      elif choice == "6":
        print("Skipping search.")
      else:
        print("Invalid choice. Skipping search.")
      
  except HttpError as err:
    print(f"An error occurred: {err}")

if __name__ == "__main__":
  main()
